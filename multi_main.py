import sys
import os
import questionary
from tqdm import tqdm

from modules.loss.hardware_module import hardware_module
from hardware_utils import hw_visualizzer, hw_test_all, hw_choose_specific, add_hw_config, remove_hw_config
from components.colors import colors
from model_utils import load_trained_model, check_model_path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def save_results_to_excel(risultati, file_name="risultati_latenza.xlsx"):
    # Crea un nuovo workbook in modalità normale (non write-only)
    wb = Workbook(write_only=False)
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title="Risultati Latenza")
    else:
        ws.title = "Risultati Latenza"

    # Aggiungi l'intestazione
    headers = ["Modello", "Nome Configurazione HW", "Latenza (s)"]
    ws.append(headers)

    # Scrivi i dati e unisci le celle per il nome del modello
    current_model = None
    start_row = 2  # La prima riga con i dati
    for i, result in enumerate(risultati, start=2):
        if result["modello"] != current_model:
            if current_model is not None:
                # Unisci le celle per il modello precedente
                ws.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
                ws.cell(start_row, 1).alignment = Alignment(vertical="center", horizontal="center")
            current_model = result["modello"]
            start_row = i
        ws.append([result["modello"], result["nome configurazione HW"], result["latenza"]])

    # Unisci le celle per l'ultimo modello
    ws.merge_cells(start_row=start_row, start_column=1, end_row=len(risultati) + 1, end_column=1)
    ws.cell(start_row, 1).alignment = Alignment(vertical="center", horizontal="center")

    # Modifica la larghezza delle colonne
    column_widths = [50, 30, 15]  # Larghezza per ogni colonna
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Salva il file
    wb.save(file_name)
    print(f"File Excel '{file_name}' creato con successo.")

def test_all_model():
    models_path = questionary.path("Enter the path of the directory with the pre-trained models").ask()
    risultati = []
    models_path = os.path.abspath(models_path)

    model_files = []

    for root, _, files in os.walk(models_path):
        for file in files:
            if file.endswith(".keras"):
                model_files.append(os.path.join(root, file))
    #model_files = [file for file in os.listdir(models_path) if file.endswith(".keras")]

    # Aggiungi la barra di caricamento
    with tqdm(total=len(model_files), desc="Testing models", unit="model") as pbar:
        for file in model_files:
            relative_path = os.path.relpath(file, models_path)
            pbar.set_postfix({"Current Model": relative_path})  # Mostra il nome del modello
            #print(f"Checking model at: {path}")
            if check_model_path(file):
                model = load_trained_model(file, show_info=False)
                for config_name, config in hw_mod.nvdla.items():
                    latency = hw_mod.get_model_latency(model, config['path'])
                    risultati.append({
                        "modello": relative_path,
                        "nome configurazione HW": config_name,
                        "latenza": latency / (10**9)
                    })
                    # print(f"Configuration: {config_name} | Latency: {latency/(10**9):.6f} secondi")
            pbar.update(1)  # Aggiorna la barra di caricamento

    save_results_to_excel(risultati, "risultati_latenza.xlsx")
    #print("File Excel 'risultati_latenza.xlsx' creato con successo.")
    """
    #Loading of the pre-trained model
    while True:
        model_path=questionary.path("Enter the path of the pre-trained model").ask()
        if not model_path:
            model_path = "gesture/dashboard/model/model.keras" #Default path if not provided
            break
        if check_model_path(model_path) == True:
            break
    model=load_trained_model(model_path)
    """


#menu for the API
print(colors.CYAN, "|  ----------- HARDWARE LATENCY TESTER ----------  |\n", colors.ENDC)
while True:
    #va bene lascaire qui il caricamento degli hardware???
    # Ensure the custom_hardware_module is correctly imported
    hw_mod=hardware_module()
    hw_names = list(hw_mod.nvdla.keys())
    choice = questionary.select(
        "Choose an option:",
        choices=[
            "1: View available hardware configurations",
            "2: Add a new hardware configuration",
            "3: Delete hardware configuration",
            "4: Test all the model",
            "5: Exit"],
    ).ask()

    if choice.startswith("1"):
        print(colors.OKBLUE, "|  ----------- AVAILABLE HARDWARE  ----------  |\n", colors.ENDC)  
        hw_visualizzer(hw_mod)
        print(colors.OKBLUE, "|  ------------------------------------  |\n", colors.ENDC)
    elif choice.startswith("2"):
        add_hw_config()
    elif choice.startswith("3"):
        remove_hw_config(hw_mod)
    elif choice.startswith("4"):
        test_all_model()
    elif choice.startswith("5"):
        sys.exit()
    else:
        print(colors.FAIL, "Invalid choice. Please try again.", colors.ENDC)

